"""PrismRAG — File adapter (Excel / CSV)."""
from __future__ import annotations

import csv
import io
from typing import BinaryIO, Iterator

from prismrag.adapters.base import Record, SourceAdapter
from prismrag.models import FileSourceConfig


class FileAdapter(SourceAdapter):
    """
    Stream records from an in-memory bytes buffer (CSV or Excel).

    The caller is responsible for reading the upload into a BytesIO before
    passing it here. Supports:
      - CSV (.csv)
      - Excel (.xlsx / .xls)  — requires openpyxl
    """

    def __init__(self, data: bytes | BinaryIO, config: FileSourceConfig):
        self._data   = data if isinstance(data, (bytes, bytearray)) else data.read()
        self._config = config
        self._rows: list[dict] | None = None

    def _load(self) -> list[dict]:
        if self._rows is not None:
            return self._rows

        fname = (self._config.filename or "").lower()
        if fname.endswith(".csv") or fname.endswith(".tsv"):
            sep = "\t" if fname.endswith(".tsv") else ","
            reader = csv.DictReader(io.StringIO(self._data.decode("utf-8-sig")), delimiter=sep)
            self._rows = list(reader)
        else:
            # Excel
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
            wb = openpyxl.load_workbook(io.BytesIO(self._data), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or "").strip() for v in row])))
            wb.close()
            self._rows = rows

        return self._rows

    def count_estimate(self) -> int | None:
        return len(self._load())

    def stream(self) -> Iterator[Record]:
        wcol = self._config.word_column
        tcol = self._config.text_column
        ccol = self._config.category_column

        for i, row in enumerate(self._load()):
            word = str(row.get(wcol) or "").strip().lower()
            text = str(row.get(tcol) or "").strip()
            if not word:
                continue
            if not text:
                text = word   # fall back to the word itself as the chunk text
            cat  = str(row.get(ccol) or "").strip() if ccol else None
            yield Record(
                word=word,
                text=text,
                ref=f"row:{i}",
                category_hint=cat or None,
                metadata={k: v for k, v in row.items() if k not in (wcol, tcol)},
            )
